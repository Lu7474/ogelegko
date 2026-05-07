import io
import json
import zipfile
from datetime import timedelta

from django.contrib.auth.models import User
from django.test import Client, TestCase
from django.utils import timezone

from .models import (
    Answer,
    Attempt,
    CatalogTask,
    CatalogTaskImage,
    ExamType,
    SchoolClass,
    Student,
    Task,
    TaskSource,
    Variant,
)
from .parser import _strip_measurement_unit, sanitize_html
from .utils import check_answer, compute_task_stats, get_grade, normalize_answer, normalize_full_name


class NormalizeAnswerTests(TestCase):
    def test_integer(self):
        self.assertEqual(normalize_answer("42"), "42")

    def test_decimal_comma(self):
        self.assertEqual(normalize_answer("3,5"), "3.5")

    def test_fraction(self):
        self.assertEqual(normalize_answer("1/2"), "0.5")

    def test_strip_spaces(self):
        self.assertEqual(normalize_answer("  7  "), "7")

    def test_empty(self):
        self.assertEqual(normalize_answer(""), "")

    def test_text(self):
        self.assertEqual(normalize_answer("abc"), "abc")

    def test_unicode_minus(self):
        # U+2212 математический минус (из MathML/ФИПИ) → ASCII дефис
        self.assertEqual(normalize_answer("\u2212" + "5"), "-5")

    def test_endash_minus(self):
        # En-dash как минус → ASCII дефис
        self.assertEqual(normalize_answer("\u2013" + "3"), "-3")

    def test_unicode_minus_in_answer_match(self):
        # Ответ с U+2212 совпадает с тем, что студент ввёл через ASCII -
        from .utils import check_answer

        self.assertTrue(check_answer("-5", "\u2212" + "5"))
        self.assertTrue(check_answer("\u2212" + "5", "-5"))


class CheckAnswerTests(TestCase):
    def test_exact(self):
        self.assertTrue(check_answer("42", "42"))

    def test_comma_vs_dot(self):
        self.assertTrue(check_answer("3,5", "3.5"))

    def test_wrong(self):
        self.assertFalse(check_answer("41", "42"))

    def test_fraction_match(self):
        self.assertTrue(check_answer("1/4", "0.25"))

    def test_empty_answer_is_wrong(self):
        self.assertFalse(check_answer("", "42"))

    def test_pipe_alternatives_first(self):
        # Ответ — первый из вариантов через |
        self.assertTrue(check_answer("234", "234|243|324"))

    def test_pipe_alternatives_last(self):
        # Ответ — последний из вариантов
        self.assertTrue(check_answer("324", "234|243|324"))

    def test_pipe_alternatives_wrong(self):
        # Ответа нет ни в одном варианте
        self.assertFalse(check_answer("999", "234|243|324"))

    def test_pi_without_symbol(self):
        # Ученик пишет 27 вместо 27π — принимается
        self.assertTrue(check_answer("27", "27π"))

    def test_pi_with_symbol(self):
        # Ученик пишет 27π — тоже принимается
        self.assertTrue(check_answer("27π", "27π"))

    def test_pi_keyword(self):
        # Ученик пишет "27pi" — принимается (pi → π)
        self.assertTrue(check_answer("27pi", "27π"))

    def test_pi_alone_not_simplified(self):
        # Ответ "π" (без коэффициента) — "1" не принимается
        self.assertFalse(check_answer("1", "π"))

    def test_pi_wrong_coeff(self):
        # Неверный коэффициент
        self.assertFalse(check_answer("28", "27π"))


class StripUnitTests(TestCase):
    """Тесты удаления единиц измерения из ответов."""

    def test_mm(self):
        self.assertEqual(_strip_measurement_unit("0.4 мм"), "0.4")

    def test_km_h(self):
        self.assertEqual(_strip_measurement_unit("60 км/ч"), "60")

    def test_percent(self):
        self.assertEqual(_strip_measurement_unit("15%"), "15")

    def test_rub(self):
        self.assertEqual(_strip_measurement_unit("1200 руб"), "1200")

    def test_no_unit(self):
        # Без единицы — возвращается без изменений
        self.assertEqual(_strip_measurement_unit("42"), "42")

    def test_text_answer(self):
        # Текст не трогается
        self.assertEqual(_strip_measurement_unit("нет"), "нет")

    def test_decimal_with_unit(self):
        self.assertEqual(_strip_measurement_unit("3,5 кг"), "3,5")


class GradeTests(TestCase):
    def test_oge_grade_5(self):
        self.assertEqual(get_grade("oge", 25), "5")

    def test_oge_grade_2(self):
        self.assertEqual(get_grade("oge", 3), "2")

    def test_ege_profile(self):
        self.assertEqual(get_grade("ege_profile", 15), "72")

    def test_ege_base_grade_3(self):
        self.assertEqual(get_grade("ege_base", 8), "3")


class AuthTests(TestCase):
    def setUp(self):
        self.client = Client()
        self.school_class = SchoolClass.objects.create(name="9А", exam_type=ExamType.OGE)
        self.student = Student(
            full_name="Тестов Тест Тестович",
            school_class=self.school_class,
        )
        self.student.set_password("test123")
        self.student.save()

    def test_login_success(self):
        resp = self.client.post(
            "/login/",
            {
                "full_name": "Тестов Тест Тестович",
                "password": "test123",
            },
        )
        self.assertEqual(resp.status_code, 302)
        self.assertIn("student_id", self.client.session)

    def test_login_wrong_password(self):
        resp = self.client.post(
            "/login/",
            {
                "full_name": "Тестов Тест Тестович",
                "password": "wrong",
            },
        )
        self.assertEqual(resp.status_code, 200)
        self.assertNotIn("student_id", self.client.session)

    def test_logout_requires_post(self):
        resp = self.client.get("/logout/")
        self.assertEqual(resp.status_code, 405)

    def test_logout_post(self):
        self.client.post(
            "/login/",
            {
                "full_name": "Тестов Тест Тестович",
                "password": "test123",
            },
        )
        resp = self.client.post("/logout/")
        self.assertEqual(resp.status_code, 302)

    def test_protected_page_redirect(self):
        resp = self.client.get("/choose/")
        self.assertEqual(resp.status_code, 302)
        self.assertIn("login", resp.url)


class ExamFlowTests(TestCase):
    def setUp(self):
        self.client = Client()
        self.school_class = SchoolClass.objects.create(name="9Б", exam_type=ExamType.OGE)
        self.student = Student(
            full_name="Экзаменов Экзамен",
            school_class=self.school_class,
        )
        self.student.set_password("pass")
        self.student.save()

        self.variant = Variant.objects.create(
            number="test001",
            exam_type=ExamType.OGE,
            max_attempts=2,
        )
        self.task1 = Task.objects.create(
            variant=self.variant,
            number=1,
            text="2+2=?",
            correct_answer="4",
            points=1,
        )
        self.task2 = Task.objects.create(
            variant=self.variant,
            number=2,
            text="3+3=?",
            correct_answer="6",
            points=1,
        )

        self.client.post(
            "/login/",
            {
                "full_name": "Экзаменов Экзамен",
                "password": "pass",
            },
        )

    def test_start_exam_creates_attempt(self):
        resp = self.client.get(f"/exam/{self.variant.id}/")
        self.assertEqual(resp.status_code, 200)
        self.assertEqual(Attempt.objects.filter(student=self.student).count(), 1)

    def test_save_answer(self):
        self.client.get(f"/exam/{self.variant.id}/")
        attempt = Attempt.objects.get(student=self.student)
        answer = Answer.objects.get(attempt=attempt, task=self.task1)

        resp = self.client.post(
            "/exam/save-answer/",
            json.dumps({"answer_id": answer.id, "value": "4"}),
            content_type="application/json",
        )
        self.assertEqual(resp.status_code, 200)
        answer.refresh_from_db()
        self.assertEqual(answer.student_answer, "4")

    def test_finish_exam_grades(self):
        self.client.get(f"/exam/{self.variant.id}/")
        attempt = Attempt.objects.get(student=self.student)

        a1 = Answer.objects.get(attempt=attempt, task=self.task1)
        a1.student_answer = "4"
        a1.save()
        a2 = Answer.objects.get(attempt=attempt, task=self.task2)
        a2.student_answer = "5"  # wrong
        a2.save()

        resp = self.client.post(f"/exam/finish/{attempt.id}/")
        self.assertEqual(resp.status_code, 302)

        attempt.refresh_from_db()
        self.assertTrue(attempt.is_finished)
        self.assertEqual(attempt.score, 1)

    def test_attempt_limit(self):
        self.client.get(f"/exam/{self.variant.id}/")
        attempt1 = Attempt.objects.get(student=self.student, is_finished=False)
        self.client.post(f"/exam/finish/{attempt1.id}/")

        self.client.get(f"/exam/{self.variant.id}/")
        attempt2 = Attempt.objects.get(student=self.student, is_finished=False)
        self.client.post(f"/exam/finish/{attempt2.id}/")

        # Третья попытка — лимит (max_attempts=2)
        resp = self.client.get(f"/exam/{self.variant.id}/")
        self.assertEqual(resp.status_code, 200)
        self.assertEqual(Attempt.objects.filter(student=self.student, is_finished=True).count(), 2)
        self.assertEqual(Attempt.objects.filter(student=self.student, is_finished=False).count(), 0)


class RetryMistakesTests(TestCase):
    """Проверяет, что «Повторить ошибки» создаёт доступный вариант."""

    def setUp(self):
        self.client = Client()
        self.school_class = SchoolClass.objects.create(name="9В", exam_type=ExamType.OGE)
        self.student = Student(full_name="Ошибков Ошибка", school_class=self.school_class)
        self.student.set_password("pass")
        self.student.save()

        self.variant = Variant.objects.create(number="retry_base", exam_type=ExamType.OGE)
        self.task = Task.objects.create(
            variant=self.variant, number=1, text="1+1=?", correct_answer="2", points=1
        )

        self.client.post("/login/", {"full_name": "Ошибков Ошибка", "password": "pass"})

        # Создаём завершённую попытку с неверным ответом
        self.client.get(f"/exam/{self.variant.id}/")
        attempt = Attempt.objects.get(student=self.student)
        answer = Answer.objects.get(attempt=attempt, task=self.task)
        answer.student_answer = "99"
        answer.save()
        self.client.post(f"/exam/finish/{attempt.id}/")
        self.attempt = Attempt.objects.get(student=self.student, is_finished=True)

    def test_retry_variant_is_active(self):
        resp = self.client.get(f"/attempt/{self.attempt.id}/retry/")
        self.assertEqual(resp.status_code, 302)

        review_number = f"ошибки_{self.variant.number}_{self.attempt.id}"
        review_variant = Variant.objects.get(number=review_number)
        self.assertTrue(review_variant.is_active)

    def test_retry_start_exam_accessible(self):
        self.client.get(f"/attempt/{self.attempt.id}/retry/")

        review_number = f"ошибки_{self.variant.number}_{self.attempt.id}"
        review_variant = Variant.objects.get(number=review_number)

        resp = self.client.get(f"/exam/{review_variant.id}/")
        self.assertEqual(resp.status_code, 200)

    def test_retry_excluded_from_random(self):
        # Если единственный активный вариант — retry-вариант, choose_variant не должен его вернуть
        self.client.get(f"/attempt/{self.attempt.id}/retry/")
        # Деактивируем оригинальный вариант
        self.variant.is_active = False
        self.variant.save()

        # Все активные варианты — только retry; случайный выбор должен вернуть ошибку «нет вариантов»
        resp = self.client.post("/choose/", {"action": "random"})
        self.assertEqual(resp.status_code, 200)
        self.assertContains(resp, "Нет доступных вариантов")


class VariantFromCatalogImageTests(TestCase):
    """Проверяет, что CatalogTaskImage копируются в TaskImage при создании варианта из каталога."""

    def setUp(self):
        self.client = Client()
        from django.contrib.auth.models import User

        self.admin = User.objects.create_user("admin2", password="admin123", is_staff=True)
        self.client.post("/admin/", {"username": "admin2", "password": "admin123"})

    def test_extra_images_copied(self):
        from django.core.files.base import ContentFile

        ct = CatalogTask.objects.create(
            task_number=1,
            exam_type=ExamType.OGE,
            text="задание",
            correct_answer="1",
            source=TaskSource.MANUAL,
        )
        # Создаём CatalogTaskImage с минимальным PNG (1×1 пиксель)
        minimal_png = (
            b"\x89PNG\r\n\x1a\n\x00\x00\x00\rIHDR\x00\x00\x00\x01\x00\x00\x00\x01"
            b"\x08\x02\x00\x00\x00\x90wS\xde\x00\x00\x00\x0cIDATx\x9cc\xf8\x0f\x00"
            b"\x00\x01\x01\x00\x05\x18\xd8N\x00\x00\x00\x00IEND\xaeB`\x82"
        )
        ci = CatalogTaskImage(task=ct, order=0)
        ci.image.save("test.png", ContentFile(minimal_png), save=True)

        resp = self.client.post(
            "/admin/variants/from-catalog/",
            {
                "variant_number": "from_catalog_img_test",
                "exam_type": "oge",
                "selected_tasks": json.dumps({"1": ct.id}),
            },
        )
        self.assertEqual(resp.status_code, 200)
        data = json.loads(resp.content)
        self.assertTrue(data.get("ok"))

        task = Task.objects.get(variant__number="from_catalog_img_test")
        self.assertEqual(task.extra_images.count(), 1)


class AdminTests(TestCase):
    def setUp(self):
        self.client = Client()
        self.admin = User.objects.create_user(
            "admin",
            password="admin123",
            is_staff=True,
        )

    def test_admin_login(self):
        resp = self.client.post(
            "/admin/",
            {
                "username": "admin",
                "password": "admin123",
            },
        )
        self.assertEqual(resp.status_code, 302)

    def test_admin_logout_requires_post(self):
        self.client.login(username="admin", password="admin123")
        resp = self.client.get("/admin/logout/")
        self.assertEqual(resp.status_code, 405)

    def test_dashboard_requires_auth(self):
        resp = self.client.get("/admin/dashboard/")
        self.assertEqual(resp.status_code, 302)

    def test_export_results(self):
        self.client.login(username="admin", password="admin123")
        resp = self.client.get("/admin/export/")
        self.assertEqual(resp.status_code, 200)
        self.assertIn("spreadsheet", resp["Content-Type"])


class NormalizeFullNameTests(TestCase):
    def test_lowercase(self):
        self.assertEqual(normalize_full_name("иванов иван иванович"), "Иванов Иван Иванович")

    def test_uppercase(self):
        self.assertEqual(normalize_full_name("ПЕТРОВ ПЁТР"), "Петров Пётр")

    def test_extra_spaces(self):
        self.assertEqual(normalize_full_name("  сидоров   сидор  "), "Сидоров Сидор")

    def test_already_normalized(self):
        self.assertEqual(normalize_full_name("Козлов Иван"), "Козлов Иван")

    def test_single_word(self):
        self.assertEqual(normalize_full_name("иван"), "Иван")

    def test_empty(self):
        self.assertEqual(normalize_full_name(""), "")

    def test_tabs_and_newlines_collapsed(self):
        self.assertEqual(normalize_full_name("иванов\tиван\nиванович"), "Иванов Иван Иванович")


class StudentLoginNormalizationTests(TestCase):
    def setUp(self):
        self.client = Client()
        self.school_class = SchoolClass.objects.create(name="10В", exam_type=ExamType.OGE)
        self.student = Student(full_name="Иванов Иван Иванович", school_class=self.school_class)
        self.student.set_password("secret")
        self.student.save()

    def test_login_lowercase_name(self):
        resp = self.client.post("/login/", {"full_name": "иванов иван иванович", "password": "secret"})
        self.assertEqual(resp.status_code, 302)
        self.assertIn("student_id", self.client.session)

    def test_login_uppercase_name(self):
        resp = self.client.post("/login/", {"full_name": "ИВАНОВ ИВАН ИВАНОВИЧ", "password": "secret"})
        self.assertEqual(resp.status_code, 302)
        self.assertIn("student_id", self.client.session)

    def test_login_extra_spaces(self):
        resp = self.client.post("/login/", {"full_name": "  Иванов   Иван  Иванович  ", "password": "secret"})
        self.assertEqual(resp.status_code, 302)
        self.assertIn("student_id", self.client.session)

    def test_login_wrong_password_still_fails(self):
        resp = self.client.post("/login/", {"full_name": "иванов иван иванович", "password": "wrong"})
        self.assertEqual(resp.status_code, 200)
        self.assertNotIn("student_id", self.client.session)


class StudentSaveNormalizationTests(TestCase):
    def setUp(self):
        self.client = Client()
        self.admin = User.objects.create_user("admin2", password="pass", is_staff=True)
        self.client.login(username="admin2", password="pass")
        self.school_class = SchoolClass.objects.create(name="11А", exam_type=ExamType.EGE_PROFILE)

    def test_student_add_normalizes_name(self):
        self.client.post(
            "/admin/students/add/",
            {
                "full_name": "петров пётр петрович",
                "password": "pwd",
                "school_class": self.school_class.id,
            },
        )
        student = Student.objects.get(school_class=self.school_class)
        self.assertEqual(student.full_name, "Петров Пётр Петрович")

    def test_student_edit_normalizes_name(self):
        student = Student(full_name="Старое Имя", school_class=self.school_class)
        student.set_password("pwd")
        student.save()
        self.client.post(
            f"/admin/students/{student.id}/edit/",
            {
                "full_name": "НОВОЕ ИМЯНОВИЧЕВ",
                "school_class": self.school_class.id,
            },
        )
        student.refresh_from_db()
        self.assertEqual(student.full_name, "Новое Имяновичев")

    def test_student_add_duplicate_case_insensitive(self):
        Student.objects.create(
            full_name="Козлов Козёл",
            school_class=self.school_class,
            password_hash="x",
        )
        resp = self.client.post(
            "/admin/students/add/",
            {
                "full_name": "козлов козёл",
                "password": "pwd",
                "school_class": self.school_class.id,
            },
        )
        self.assertEqual(resp.status_code, 200)
        self.assertEqual(Student.objects.filter(school_class=self.school_class).count(), 1)

    def test_excel_import_normalizes_name(self):
        import io

        import openpyxl

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["ФИО", "Пароль", "Класс"])
        ws.append(["сидоров сидор", "pass1", "11А"])
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        buf.name = "students.xlsx"
        self.client.post("/admin/students/import/", {"file": buf})
        student = Student.objects.get(school_class=self.school_class)
        self.assertEqual(student.full_name, "Сидоров Сидор")

    def test_excel_import_duplicate_in_file(self):
        import io

        import openpyxl

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["ФИО", "Пароль", "Класс"])
        ws.append(["Волков Волк", "p1", "11А"])
        ws.append(["волков волк", "p2", "11А"])
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        buf.name = "students.xlsx"
        resp = self.client.post("/admin/students/import/", {"file": buf})
        self.assertEqual(Student.objects.filter(school_class=self.school_class).count(), 1)
        self.assertContains(resp, "дубль")

    def test_excel_import_duplicate_in_db(self):
        import io

        import openpyxl

        Student.objects.create(
            full_name="Зайцев Заяц",
            school_class=self.school_class,
            password_hash="x",
        )
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["ФИО", "Пароль", "Класс"])
        ws.append(["зайцев заяц", "p1", "11А"])
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        buf.name = "students.xlsx"
        resp = self.client.post("/admin/students/import/", {"file": buf})
        self.assertEqual(Student.objects.filter(school_class=self.school_class).count(), 1)
        self.assertContains(resp, "уже существует")


class SecurityTests(TestCase):
    """Тесты безопасности: изоляция данных, санитизация, лимиты."""

    def setUp(self):
        self.client = Client()
        self.school_class = SchoolClass.objects.create(name="10Б", exam_type=ExamType.EGE_PROFILE)
        self.student1 = Student(full_name="Тест Один", school_class=self.school_class)
        self.student1.set_password("pass1")
        self.student1.save()
        self.student2 = Student(full_name="Тест Два", school_class=self.school_class)
        self.student2.set_password("pass2")
        self.student2.save()
        self.variant = Variant.objects.create(number="sec_v1", exam_type=ExamType.EGE_PROFILE)
        self.task = Task.objects.create(
            variant=self.variant,
            number="1",
            text="Задание",
            correct_answer="42",
        )

    def _login(self, student):
        session = self.client.session
        session["student_id"] = student.id
        session.save()
        student.session_key = self.client.session.session_key
        student.save(update_fields=["session_key"])

    def test_student_cannot_access_other_results(self):
        """Ученик не должен видеть результаты другого ученика."""
        attempt = Attempt.objects.create(
            student=self.student2,
            variant=self.variant,
            is_finished=True,
            score=1,
            max_score=1,
        )
        self._login(self.student1)
        resp = self.client.get(f"/results/{attempt.id}/")
        self.assertNotEqual(resp.status_code, 200)

    def test_sanitize_html_strips_script_tags(self):
        """sanitize_html удаляет script-теги."""
        xss = '<script>alert("xss")</script>текст задания'
        result = sanitize_html(xss)
        self.assertNotIn("<script>", result)
        self.assertIn("текст задания", result)

    def test_sanitize_html_strips_event_handlers(self):
        """sanitize_html удаляет атрибуты-обработчики событий."""
        xss = '<img src="x" onerror="alert(1)">'
        result = sanitize_html(xss)
        self.assertNotIn("onerror", result)

    def test_attempt_limit_enforced(self):
        """Ученик не может создать попытку сверх лимита."""
        self.variant.max_attempts = 1
        self.variant.save()
        Attempt.objects.create(
            student=self.student1,
            variant=self.variant,
            is_finished=True,
            score=0,
            max_score=1,
        )
        self._login(self.student1)
        resp = self.client.get(f"/start/{self.variant.id}/")
        self.assertNotEqual(resp.status_code, 302)
        self.assertEqual(Attempt.objects.filter(student=self.student1, is_finished=False).count(), 0)

    def test_pdf_size_limit_rejected(self):
        """PDF больше 50 МБ отклоняется без записи на диск."""
        User.objects.create_user("admin_sec", password="pass", is_staff=True)
        self.client.login(username="admin_sec", password="pass")
        big_file = io.BytesIO(b"0" * (51 * 1024 * 1024))
        big_file.name = "big.pdf"
        big_file.size = 51 * 1024 * 1024
        resp = self.client.post(
            "/admin/catalog/import-pdf/",
            {
                "exam_type": ExamType.EGE_PROFILE,
                "pdf_files": big_file,
                "mode": "catalog",
                "format": "print_solve",
            },
        )
        self.assertContains(resp, "слишком большой")

    def test_compute_task_stats_empty(self):
        """compute_task_stats возвращает пустой dict для пустого queryset."""
        result = compute_task_stats(Answer.objects.none())
        self.assertEqual(result, {})

    def test_sanitize_html_onerror_unquoted(self):
        result = sanitize_html("<img src=x onerror=alert(1)>")
        self.assertNotIn("onerror", result)

    def test_sanitize_html_javascript_href(self):
        result = sanitize_html('<a href="javascript:alert(1)">click</a>')
        self.assertNotIn("javascript:", result)

    def test_zip_import_text_not_sanitized(self):
        """ZIP-импорт не вызывает sanitize_html — документируем текущее поведение."""
        from exam.services.variant_archive import import_variants_from_zip

        xss = "<img src=x onerror=alert(1)>"
        folder = "xss_zip"
        manifest = {
            "format_version": 1,
            "exported_at": "2026-01-01T00:00:00+00:00",
            "app": "EGE",
            "variants_count": 1,
            "variants": [{"folder": folder, "number": "xss_zip_v", "exam_type": "oge"}],
        }
        vdata = {
            "format_version": 1,
            "number": "xss_zip_v",
            "exam_type": "oge",
            "max_attempts": 1,
            "is_active": True,
            "tasks": [{"folder": "tasks/1", "number": "1"}],
        }
        tdata = {
            "format_version": 1,
            "number": "1",
            "text": xss,
            "correct_answer": "1",
            "source": "manual",
            "points": 1,
            "manual_grading": False,
            "no_student_input": False,
            "shared_context": "",
            "image": None,
            "shared_context_image": None,
            "extra_images": [],
        }
        buf = io.BytesIO()
        with zipfile.ZipFile(buf, "w") as zf:
            zf.writestr("manifest.json", json.dumps(manifest))
            zf.writestr(f"{folder}/variant.json", json.dumps(vdata))
            zf.writestr(f"{folder}/tasks/1/task.json", json.dumps(tdata))
        buf.seek(0)

        class FakeFile:
            def read(self):
                return buf.read()

        import_variants_from_zip(FakeFile())
        task = Task.objects.get(variant__number="xss_zip_v")
        # Текущее поведение: текст сохраняется без санитизации (известная уязвимость)
        self.assertEqual(task.text, xss)


class ExamTimerTests(TestCase):
    def setUp(self):
        self.client = Client()
        self.school_class = SchoolClass.objects.create(name="10Г", exam_type=ExamType.OGE)
        self.student = Student(full_name="Таймеров Тест", school_class=self.school_class)
        self.student.set_password("pass")
        self.student.save()
        self.variant = Variant.objects.create(number="timer_v1", exam_type=ExamType.OGE)
        self.task = Task.objects.create(
            variant=self.variant, number="1", text="1+1=?", correct_answer="2", points=1
        )
        self.client.post("/login/", {"full_name": "Таймеров Тест", "password": "pass"})

    def _expire_attempt(self, attempt):
        Attempt.objects.filter(id=attempt.id).update(started_at=timezone.now() - timedelta(hours=5))
        attempt.refresh_from_db()

    def test_save_answer_returns_403_when_expired(self):
        self.client.get(f"/exam/{self.variant.id}/")
        attempt = Attempt.objects.get(student=self.student, is_finished=False)
        answer = Answer.objects.get(attempt=attempt, task=self.task)
        self._expire_attempt(attempt)

        resp = self.client.post(
            "/exam/save-answer/",
            json.dumps({"answer_id": answer.id, "value": "2"}),
            content_type="application/json",
        )
        self.assertEqual(resp.status_code, 403)
        self.assertEqual(json.loads(resp.content).get("error"), "time_expired")

    def test_start_exam_redirects_to_results_when_expired(self):
        self.client.get(f"/exam/{self.variant.id}/")
        attempt = Attempt.objects.get(student=self.student, is_finished=False)
        self._expire_attempt(attempt)

        resp = self.client.get(f"/exam/{self.variant.id}/")
        self.assertEqual(resp.status_code, 302)
        self.assertIn("results", resp.url)
        attempt.refresh_from_db()
        self.assertTrue(attempt.is_finished)
