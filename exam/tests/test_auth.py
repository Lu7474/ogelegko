import io

import openpyxl
from django.contrib.auth.models import User
from django.test import Client, TestCase

from ..models import ExamType, SchoolClass, Student


class AuthTests(TestCase):
    def setUp(self):
        self.client = Client()
        self.school_class = SchoolClass.objects.create(name="9А", exam_type=ExamType.OGE)
        self.student = Student(
            full_name="Тестов Тест Тестович",
            school_class=self.school_class,
        )
        self.student.set_password("test123")
        self.student.save()

    def test_login_success(self):
        resp = self.client.post(
            "/login/",
            {
                "full_name": "Тестов Тест Тестович",
                "password": "test123",
            },
        )
        self.assertEqual(resp.status_code, 302)
        self.assertIn("student_id", self.client.session)

    def test_login_wrong_password(self):
        resp = self.client.post(
            "/login/",
            {
                "full_name": "Тестов Тест Тестович",
                "password": "wrong",
            },
        )
        self.assertEqual(resp.status_code, 200)
        self.assertNotIn("student_id", self.client.session)

    def test_logout_requires_post(self):
        resp = self.client.get("/logout/")
        self.assertEqual(resp.status_code, 405)

    def test_logout_post(self):
        self.client.post(
            "/login/",
            {
                "full_name": "Тестов Тест Тестович",
                "password": "test123",
            },
        )
        resp = self.client.post("/logout/")
        self.assertEqual(resp.status_code, 302)

    def test_protected_page_redirect(self):
        resp = self.client.get("/choose/")
        self.assertEqual(resp.status_code, 302)
        self.assertIn("login", resp.url)


class StudentLoginNormalizationTests(TestCase):
    def setUp(self):
        self.client = Client()
        self.school_class = SchoolClass.objects.create(name="10В", exam_type=ExamType.OGE)
        self.student = Student(full_name="Иванов Иван Иванович", school_class=self.school_class)
        self.student.set_password("secret")
        self.student.save()

    def test_login_lowercase_name(self):
        resp = self.client.post("/login/", {"full_name": "иванов иван иванович", "password": "secret"})
        self.assertEqual(resp.status_code, 302)
        self.assertIn("student_id", self.client.session)

    def test_login_uppercase_name(self):
        resp = self.client.post("/login/", {"full_name": "ИВАНОВ ИВАН ИВАНОВИЧ", "password": "secret"})
        self.assertEqual(resp.status_code, 302)
        self.assertIn("student_id", self.client.session)

    def test_login_extra_spaces(self):
        resp = self.client.post("/login/", {"full_name": "  Иванов   Иван  Иванович  ", "password": "secret"})
        self.assertEqual(resp.status_code, 302)
        self.assertIn("student_id", self.client.session)

    def test_login_wrong_password_still_fails(self):
        resp = self.client.post("/login/", {"full_name": "иванов иван иванович", "password": "wrong"})
        self.assertEqual(resp.status_code, 200)
        self.assertNotIn("student_id", self.client.session)


class StudentSaveNormalizationTests(TestCase):
    def setUp(self):
        self.client = Client()
        self.admin = User.objects.create_user("admin2", password="pass", is_staff=True)
        self.client.login(username="admin2", password="pass")
        self.school_class = SchoolClass.objects.create(name="11А", exam_type=ExamType.EGE_PROFILE)

    def test_student_add_normalizes_name(self):
        self.client.post(
            "/admin/students/add/",
            {
                "full_name": "петров пётр петрович",
                "password": "pwd",
                "school_class": self.school_class.id,
            },
        )
        student = Student.objects.get(school_class=self.school_class)
        self.assertEqual(student.full_name, "Петров Пётр Петрович")

    def test_student_edit_normalizes_name(self):
        student = Student(full_name="Старое Имя", school_class=self.school_class)
        student.set_password("pwd")
        student.save()
        self.client.post(
            f"/admin/students/{student.id}/edit/",
            {
                "full_name": "НОВОЕ ИМЯНОВИЧЕВ",
                "school_class": self.school_class.id,
            },
        )
        student.refresh_from_db()
        self.assertEqual(student.full_name, "Новое Имяновичев")

    def test_student_add_duplicate_case_insensitive(self):
        Student.objects.create(
            full_name="Козлов Козёл",
            school_class=self.school_class,
            password_hash="x",
        )
        resp = self.client.post(
            "/admin/students/add/",
            {
                "full_name": "козлов козёл",
                "password": "pwd",
                "school_class": self.school_class.id,
            },
        )
        self.assertEqual(resp.status_code, 200)
        self.assertEqual(Student.objects.filter(school_class=self.school_class).count(), 1)

    def test_excel_import_normalizes_name(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["ФИО", "Пароль", "Класс"])
        ws.append(["сидоров сидор", "pass1", "11А"])
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        buf.name = "students.xlsx"
        self.client.post("/admin/students/import/", {"file": buf})
        student = Student.objects.get(school_class=self.school_class)
        self.assertEqual(student.full_name, "Сидоров Сидор")

    def test_excel_import_duplicate_in_file(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["ФИО", "Пароль", "Класс"])
        ws.append(["Волков Волк", "p1", "11А"])
        ws.append(["волков волк", "p2", "11А"])
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        buf.name = "students.xlsx"
        resp = self.client.post("/admin/students/import/", {"file": buf})
        self.assertEqual(Student.objects.filter(school_class=self.school_class).count(), 1)
        self.assertContains(resp, "дубль")

    def test_excel_import_duplicate_in_db(self):
        Student.objects.create(
            full_name="Зайцев Заяц",
            school_class=self.school_class,
            password_hash="x",
        )
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["ФИО", "Пароль", "Класс"])
        ws.append(["зайцев заяц", "p1", "11А"])
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        buf.name = "students.xlsx"
        resp = self.client.post("/admin/students/import/", {"file": buf})
        self.assertEqual(Student.objects.filter(school_class=self.school_class).count(), 1)
        self.assertContains(resp, "уже существует")
