import json
import logging
from collections import defaultdict
from datetime import timedelta

from django.contrib.auth.hashers import make_password
from django.db import IntegrityError
from django.db.models import Count, Q
from django.http import HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.utils import timezone
from django.views.decorators.http import require_POST

from .admin_views import _paginate, _safe_int, admin_required
from .models import Answer, Attempt, ExamType, SchoolClass, Student, Task, Variant

logger = logging.getLogger(__name__)


# ===== КЛАССЫ =====


@admin_required
def class_list(request):
    classes = SchoolClass.objects.annotate(student_count=Count("students"))
    return render(request, "admin/classes.html", {"classes": classes})


@admin_required
def class_add(request):
    error = ""
    if request.method == "POST":
        name = request.POST.get("name", "").strip()
        exam_type = request.POST.get("exam_type", "")
        if not name:
            error = "Введите название класса"
        elif exam_type not in dict(ExamType.choices):
            error = "Выберите тип экзамена"
        else:
            try:
                SchoolClass.objects.create(name=name, exam_type=exam_type)
                return redirect("admin_classes")
            except IntegrityError:
                error = f"Класс '{name}' уже существует"

    return render(
        request,
        "admin/class_form.html",
        {
            "exam_types": ExamType.choices,
            "error": error,
        },
    )


@admin_required
def class_edit(request, class_id):
    school_class = get_object_or_404(SchoolClass, id=class_id)
    error = ""
    if request.method == "POST":
        name = request.POST.get("name", "").strip()
        exam_type = request.POST.get("exam_type", "")
        if not name:
            error = "Введите название класса"
        elif exam_type not in dict(ExamType.choices):
            error = "Выберите тип экзамена"
        else:
            try:
                school_class.name = name
                school_class.exam_type = exam_type
                school_class.save()
                return redirect("admin_classes")
            except IntegrityError:
                error = f"Класс '{name}' уже существует"

    return render(
        request,
        "admin/class_form.html",
        {
            "school_class": school_class,
            "exam_types": ExamType.choices,
            "error": error,
        },
    )


@admin_required
@require_POST
def class_delete(request, class_id):
    school_class = get_object_or_404(SchoolClass, id=class_id)
    school_class.delete()
    return redirect("admin_classes")


@admin_required
@require_POST
def class_toggle(request, class_id):
    school_class = get_object_or_404(SchoolClass, id=class_id)
    school_class.is_active = not school_class.is_active
    school_class.save(update_fields=["is_active"])
    return redirect("admin_classes")


@admin_required
def class_stats(request, class_id):
    school_class = get_object_or_404(SchoolClass, id=class_id)
    students = list(school_class.students.all())

    # Одним запросом получаем все завершённые попытки класса с кол-вом верных ответов
    all_attempts = list(
        Attempt.objects.filter(student__school_class=school_class, is_finished=True)
        .annotate(
            correct_count_db=Count("answers", filter=Q(answers__is_correct=True)),
            total_count_db=Count("variant__tasks", distinct=True),
        )
        .select_related("variant", "student")
        .order_by("-finished_at")
    )

    attempts_by_student = defaultdict(list)
    for attempt in all_attempts:
        attempts_by_student[attempt.student_id].append(attempt)

    student_stats_list = []
    total_attempts = 0
    all_percentages = []

    for student in students:
        student_attempts = attempts_by_student[student.id]
        count = len(student_attempts)
        total_attempts += count

        if count > 0:
            percentages = [
                round(a.correct_count_db / a.total_count_db * 100) if a.total_count_db else 0
                for a in student_attempts
            ]
            avg_pct = round(sum(percentages) / len(percentages))
            all_percentages.extend(percentages)
            last_attempt = student_attempts[0]
        else:
            avg_pct = None
            last_attempt = None

        student_stats_list.append(
            {
                "student": student,
                "attempts_count": count,
                "avg_percentage": avg_pct,
                "last_attempt": last_attempt,
            }
        )

    class_avg = round(sum(all_percentages) / len(all_percentages)) if all_percentages else None

    # Аналитика по номерам заданий для всего класса
    class_task_stats = {}
    for answer in (
        Answer.objects.filter(attempt__student__school_class=school_class, attempt__is_finished=True)
        .select_related("task")
        .exclude(task__isnull=True)
    ):
        num = answer.task.number
        if num not in class_task_stats:
            class_task_stats[num] = {"tried": 0, "correct": 0}
        class_task_stats[num]["tried"] += 1
        if answer.is_correct is True:
            class_task_stats[num]["correct"] += 1

    def _sort_key(n):
        try:
            return (0, int(str(n).split(".")[0]), str(n))
        except (ValueError, IndexError):
            return (1, 0, str(n))

    class_task_perf = [
        {
            "number": num,
            "tried": data["tried"],
            "correct": data["correct"],
            "percentage": round(data["correct"] / data["tried"] * 100) if data["tried"] else 0,
        }
        for num, data in sorted(class_task_stats.items(), key=lambda x: _sort_key(x[0]))
    ]

    return render(
        request,
        "admin/class_stats.html",
        {
            "school_class": school_class,
            "student_stats": student_stats_list,
            "total_attempts": total_attempts,
            "class_avg": class_avg,
            "class_task_perf": class_task_perf,
        },
    )


@admin_required
def class_stats_excel(request, class_id):
    """Экспорт аналитики класса в Excel."""
    try:
        import openpyxl
        from openpyxl.styles import Alignment, Font, PatternFill
    except ImportError:
        return HttpResponse("openpyxl не установлен", status=500)

    school_class = get_object_or_404(SchoolClass, id=class_id)

    class_task_stats = {}
    for answer in (
        Answer.objects.filter(attempt__student__school_class=school_class, attempt__is_finished=True)
        .select_related("task")
        .exclude(task__isnull=True)
    ):
        num = answer.task.number
        if num not in class_task_stats:
            class_task_stats[num] = {"tried": 0, "correct": 0}
        class_task_stats[num]["tried"] += 1
        if answer.is_correct is True:
            class_task_stats[num]["correct"] += 1

    def _sort_key(n):
        try:
            return (0, int(str(n).split(".")[0]), str(n))
        except (ValueError, IndexError):
            return (1, 0, str(n))

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Аналитика класса"

    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    center = Alignment(horizontal="center")

    headers = ["Номер задания", "Попыток", "Верно", "%"]
    ws.append(headers)
    for col, _ in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center

    for num, data in sorted(class_task_stats.items(), key=lambda x: _sort_key(x[0])):
        pct = round(data["correct"] / data["tried"] * 100) if data["tried"] else 0
        ws.append([num, data["tried"], data["correct"], pct])
        if pct < 50:
            for col in range(1, 5):
                ws.cell(row=ws.max_row, column=col).fill = PatternFill(
                    start_color="FFCCCC", end_color="FFCCCC", fill_type="solid"
                )

    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 18

    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = f'attachment; filename="analytics_{school_class.name}.xlsx"'
    wb.save(response)
    return response


# ===== УЧЕНИКИ =====


@admin_required
def student_list(request):
    class_filter = request.GET.get("class", "")
    has_pending = request.GET.get("has_pending", "")
    students = Student.objects.select_related("school_class").annotate(
        attempts_count=Count("attempts", filter=Q(attempts__is_finished=True))
    )
    if class_filter:
        filter_id = _safe_int(class_filter)
        if filter_id:
            students = students.filter(school_class_id=filter_id)
    if has_pending:
        students = students.filter(
            attempts__is_finished=True,
            attempts__answers__is_correct__isnull=True,
        ).distinct()

    classes = SchoolClass.objects.all()
    return render(
        request,
        "admin/students.html",
        {
            "students": students,
            "classes": classes,
            "class_filter": class_filter,
            "has_pending": has_pending,
        },
    )


@admin_required
def student_add(request):
    error = ""
    if request.method == "POST":
        full_name = request.POST.get("full_name", "").strip()
        password = request.POST.get("password", "").strip()
        class_id = _safe_int(request.POST.get("school_class", ""))

        if not full_name:
            error = "Введите ФИО"
        elif not password:
            error = "Введите пароль"
        elif not class_id:
            error = "Выберите класс"
        else:
            try:
                school_class = SchoolClass.objects.get(id=class_id)
                student = Student(full_name=full_name, school_class=school_class)
                student.set_password(password)
                student.save()
                return redirect("admin_students")
            except SchoolClass.DoesNotExist:
                error = "Выбранный класс не найден"
            except IntegrityError:
                error = f"Ученик '{full_name}' уже существует"

    classes = SchoolClass.objects.all()
    return render(request, "admin/student_form.html", {"classes": classes, "error": error})


@admin_required
def student_edit(request, student_id):
    student = get_object_or_404(Student, id=student_id)
    error = ""
    if request.method == "POST":
        full_name = request.POST.get("full_name", "").strip()
        class_id = _safe_int(request.POST.get("school_class", ""))
        password = request.POST.get("password", "").strip()

        if not full_name:
            error = "Введите ФИО"
        elif not class_id:
            error = "Выберите класс"
        else:
            try:
                school_class = SchoolClass.objects.get(id=class_id)
                student.full_name = full_name
                student.school_class = school_class
                if password:
                    student.set_password(password)
                student.save()
                return redirect("admin_students")
            except SchoolClass.DoesNotExist:
                error = "Выбранный класс не найден"
            except IntegrityError:
                error = f"Ученик '{full_name}' уже существует"

    classes = SchoolClass.objects.all()
    return render(
        request,
        "admin/student_form.html",
        {
            "student": student,
            "classes": classes,
            "error": error,
        },
    )


@admin_required
@require_POST
def student_delete(request, student_id):
    student = get_object_or_404(Student, id=student_id)
    student.delete()
    return redirect("admin_students")


@admin_required
def student_import(request):
    """Массовый импорт учеников из Excel."""
    errors = []
    success_count = 0

    if request.method == "POST" and request.FILES.get("file"):
        try:
            import openpyxl
        except ImportError:
            errors.append("Библиотека openpyxl не установлена. Выполните: pip install openpyxl")
            return render(
                request,
                "admin/student_import.html",
                {"errors": errors, "success_count": success_count},
            )

        uploaded_file = request.FILES["file"]
        if not uploaded_file.name.endswith(".xlsx"):
            errors.append("Поддерживается только формат .xlsx")
        elif uploaded_file.size > 5 * 1024 * 1024:
            errors.append("Файл слишком большой (максимум 5 МБ)")
        else:
            try:
                wb = openpyxl.load_workbook(uploaded_file)
                ws = wb.active

                seen_in_file = {}  # (class_name, full_name) -> row_num
                for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                    if not row or not row[0]:
                        continue
                    try:
                        if len(row) < 3 or not row[1] or not row[2]:
                            errors.append(
                                f"Строка {row_num}: не все поля заполнены (нужно: ФИО, пароль, класс)"
                            )
                            continue

                        full_name = str(row[0]).strip()
                        password = str(row[1]).strip()
                        class_name = str(row[2]).strip()

                        dup_key = (class_name, full_name)
                        if dup_key in seen_in_file:
                            errors.append(
                                f"Строка {row_num}: дубль — '{full_name}' ({class_name}) "
                                f"уже есть в строке {seen_in_file[dup_key]}"
                            )
                            continue
                        seen_in_file[dup_key] = row_num

                        school_class = SchoolClass.objects.get(name=class_name)
                        student = Student(full_name=full_name, school_class=school_class)
                        student.set_password(password)
                        student.save()
                        success_count += 1
                    except SchoolClass.DoesNotExist:
                        errors.append(f"Строка {row_num}: класс '{class_name}' не найден")
                    except Exception as e:
                        logger.exception("Ошибка импорта строки %d", row_num)
                        errors.append(f"Строка {row_num}: {str(e)}")

            except Exception as e:
                logger.exception("Ошибка чтения Excel файла")
                errors.append(f"Ошибка чтения файла: {str(e)}")

    return render(
        request,
        "admin/student_import.html",
        {"errors": errors, "success_count": success_count},
    )


@admin_required
def student_stats(request, student_id):
    student = get_object_or_404(Student.objects.select_related("school_class"), id=student_id)
    attempts = list(
        Attempt.objects.filter(student=student, is_finished=True)
        .annotate(
            correct_count_db=Count("answers", filter=Q(answers__is_correct=True)),
            total_count_db=Count("variant__tasks", distinct=True),
        )
        .select_related("variant")
        .order_by("-finished_at")
    )

    avg_percentage = None
    if attempts:
        percentages = [
            round(a.correct_count_db / a.total_count_db * 100) if a.total_count_db else 0 for a in attempts
        ]
        avg_percentage = round(sum(percentages) / len(percentages))

    chart_data = [
        {
            "date": a.finished_at.strftime("%d.%m.%Y"),
            "variant": a.variant.number,
            "score": a.score,
            "max_score": a.max_score,
            "percentage": round(a.correct_count_db / a.total_count_db * 100) if a.total_count_db else 0,
        }
        for a in reversed(attempts)
    ]

    # Аналитика по номерам заданий
    task_stats = {}
    for answer in (
        Answer.objects.filter(attempt__student=student, attempt__is_finished=True)
        .select_related("task")
        .exclude(task__isnull=True)
    ):
        num = answer.task.number
        if num not in task_stats:
            task_stats[num] = {"tried": 0, "correct": 0}
        task_stats[num]["tried"] += 1
        if answer.is_correct is True:
            task_stats[num]["correct"] += 1

    def _sort_key(n):
        try:
            return (0, int(str(n).split(".")[0]), str(n))
        except (ValueError, IndexError):
            return (1, 0, str(n))

    task_perf_list = [
        {
            "number": num,
            "tried": data["tried"],
            "correct": data["correct"],
            "percentage": round(data["correct"] / data["tried"] * 100) if data["tried"] else 0,
        }
        for num, data in sorted(task_stats.items(), key=lambda x: _sort_key(x[0]))
    ]
    weak_tasks = [t for t in task_perf_list if t["percentage"] < 50 and t["tried"] > 0]

    return render(
        request,
        "admin/student_stats.html",
        {
            "student": student,
            "attempts": attempts,
            "avg_percentage": avg_percentage,
            "chart_data": json.dumps(chart_data),
            "task_perf_list": task_perf_list,
            "weak_tasks": weak_tasks,
        },
    )


# ===== ПОПЫТКИ =====


@admin_required
def attempt_detail(request, attempt_id):
    attempt = get_object_or_404(Attempt, id=attempt_id, is_finished=True)
    answers = attempt.answers.select_related("task").order_by("task__id")
    has_pending = answers.filter(is_correct=None).exists()
    return render(
        request,
        "admin/attempt_detail.html",
        {
            "attempt": attempt,
            "answers": answers,
            "has_pending": has_pending,
        },
    )


@admin_required
@require_POST
def attempt_grade_answer(request, answer_id):
    from .views import _recalculate_attempt_score

    answer = get_object_or_404(Answer, id=answer_id, task__manual_grading=True)

    if "points" in request.POST:
        try:
            pts = int(request.POST["points"])
            pts = max(0, min(pts, answer.task.points))
            answer.awarded_points = pts
            answer.is_correct = pts > 0
        except (ValueError, TypeError):
            pass
    else:
        value = request.POST.get("is_correct")
        if value == "true":
            answer.is_correct = True
            answer.awarded_points = answer.task.points
        elif value == "false":
            answer.is_correct = False
            answer.awarded_points = 0
        else:
            answer.is_correct = None
            answer.awarded_points = None

    answer.save(update_fields=["is_correct", "awarded_points"])
    _recalculate_attempt_score(answer.attempt)
    return redirect("admin_attempt_detail", attempt_id=answer.attempt_id)


@admin_required
def api_new_attempts(request):
    """JSON: попытки завершённые после переданного ISO-timestamp (или за последние 24ч)."""
    from datetime import datetime

    from django.urls import reverse

    since_str = request.GET.get("since", "")
    try:
        since_dt = datetime.fromisoformat(since_str.replace("Z", "+00:00"))
        if timezone.is_naive(since_dt):
            since_dt = timezone.make_aware(since_dt)
    except (ValueError, TypeError, AttributeError):
        since_dt = timezone.now() - timedelta(hours=24)

    attempts = (
        Attempt.objects.filter(is_finished=True, finished_at__gt=since_dt)
        .select_related("student", "variant", "student__school_class")
        .order_by("-finished_at")[:20]
    )

    data = {
        "count": attempts.count(),
        "attempts": [
            {
                "student": a.student.full_name,
                "class": a.student.school_class.name,
                "variant": a.variant.number,
                "finished_at": a.finished_at.strftime("%d.%m %H:%M"),
                "grade": a.grade,
                "url": reverse("admin_attempt_detail", args=[a.id]),
            }
            for a in attempts
        ],
    }
    return JsonResponse(data)


@admin_required
@require_POST
def attempt_delete(request, attempt_id):
    attempt = get_object_or_404(Attempt, id=attempt_id)
    student_id = attempt.student_id
    attempt.delete()
    return redirect("admin_student_stats", student_id=student_id)
